# -*- coding: utf-8 -*-
"""从桌面题库 xlsx 生成 static-site/js/data.js（权威数据源）"""
import glob
import io
import json
import os

import openpyxl

PER_LEVEL = 5  # 每关 5 题


def find_xlsx():
    desktop = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop")
    if not os.path.isdir(desktop):
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    files = glob.glob(os.path.join(desktop, "*.xlsx")) + glob.glob(os.path.join(desktop, "*.xls"))
    if not files:
        raise SystemExit("未在桌面找到 xlsx 题库文件")
    return files[0]


def main():
    path = find_xlsx()
    print("解析:", path)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    print("表头:", header)
    print("数据行数:", len(rows) - 1)

    questions = []
    for idx, row in enumerate(rows[1:], start=1):
        # 列：序号 题型 分类 题目 选项A 选项B 选项C 答案
        no, qtype, category, text, opt_a, opt_b, opt_c, answer = (list(row) + [None] * 8)[:8]
        if text is None:
            continue
        text = str(text).strip()
        qtype = "" if qtype is None else str(qtype).strip()
        category = "" if category is None else str(category).strip()
        answer = "" if answer is None else str(answer).strip()
        # 判断题：xlsx 无选项，答案列为 √/× → 转为「正确/错误」选项 + A/B
        if qtype == "判断题":
            opt_a, opt_b, opt_c = "正确", "错误", ""
            answer = "A" if "√" in answer else "B"
            explanation = "√ 正确 / × 错误"
        else:
            opt_a = "" if opt_a is None else str(opt_a).strip()
            opt_b = "" if opt_b is None else str(opt_b).strip()
            opt_c = "" if opt_c is None else str(opt_c).strip()
            answer = answer.upper()
            explanation = ""
        level = (idx - 1) // PER_LEVEL + 1
        questions.append({
            "id": idx,
            "level": level,
            "category": category,
            "text": text,
            "option_a": opt_a,
            "option_b": opt_b,
            "option_c": opt_c,
            "option_d": "",
            "answer": answer,
            "explanation": explanation,
        })

    total_levels = (len(questions) + PER_LEVEL - 1) // PER_LEVEL
    data = {"per_level": PER_LEVEL, "total_levels": total_levels, "questions": questions}

    out = "static-site/js/data.js"
    with io.open(out, "w", encoding="utf-8", newline="\n") as f:
        f.write("window.QUIZ_DATA = ")
        f.write(json.dumps(data, ensure_ascii=False, indent=1))
        f.write(";\n")
    print("已生成:", out, "| 题数:", len(questions), "| 关卡:", total_levels)


if __name__ == "__main__":
    main()
